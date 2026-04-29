"""
CDE 受理品种目录 数据采集

抓取 国家药品监督管理局药品审评中心 (cde.org.cn) 的"受理品种目录"首页数据。
目标表格: 受理品种目录，每页 10 条记录，8 列字段。

注意事项:
  - CDE 部署了 CloudWAF（网宿），headless 模式会被屏蔽
  - 必须使用非 headless Chrome，等待 WAF JS 挑战通过（~20-30s）
  - 数据在 LayUI 模板引擎渲染的 tbody#acceptVarietyInfoTbody 中

运行:
  python3 scrape_cde.py
"""

import json
import os
import shutil
import sys
import time
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# ── 配置 ──────────────────────────────────────────────────────
URL = "https://www.cde.org.cn/main/xxgk/listpage/9f9c74c73e0f8f56a8bfbc646055026d"

# 输出文件路径（相对于脚本所在目录）
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_PATH = SCRIPT_DIR / "cde_data.json"
OUTPUT_XLSX_PATH = SCRIPT_DIR / "cde_data.xlsx"

# 等待超时（秒）
WAF_WAIT_SECONDS = 30   # WAF JS 挑战最大等待时间
DATA_WAIT_SECONDS = 15   # 数据渲染最大等待时间

# 表格列的 CSS 列名
COLUMNS = ["序号", "受理号", "药品名称", "药品类型",
           "申请类型", "注册分类", "企业名称", "承办日期"]


# ── ChromeDriver 发现 ─────────────────────────────────────────

def find_chromedriver() -> Optional[str]:
    """自动查找 chromedriver，返回可执行文件路径或 None。
    按以下优先级查找：
    1. 环境变量 CHROMEDRIVER_PATH
    2. 系统 PATH 中的 chromedriver
    3. webdriver-manager 缓存 (~/.wdm/)
    4. Homebrew 安装路径
    """
    # 1. 环境变量
    env_path = os.environ.get("CHROMEDRIVER_PATH")
    if env_path and Path(env_path).is_file():
        return env_path

    # 2. 系统 PATH
    path = shutil.which("chromedriver")
    if path:
        return path

    # 3. webdriver-manager 缓存
    wdm_root = Path.home() / ".wdm" / "drivers" / "chromedriver"
    if wdm_root.exists():
        # 遍历所有版本的缓存目录
        for driver_file in wdm_root.rglob("chromedriver*"):
            if driver_file.is_file() and os.access(driver_file, os.X_OK):
                return str(driver_file)

    # 4. Homebrew
    brew_paths = [
        Path("/opt/homebrew/bin/chromedriver"),
        Path("/usr/local/bin/chromedriver"),
    ]
    for p in brew_paths:
        if p.is_file():
            return str(p)

    return None


# ── 浏览器启动 ────────────────────────────────────────────────

def create_driver(chromedriver_path: Optional[str]) -> webdriver.Chrome:
    """创建非 headless Chrome WebDriver 实例。"""
    options = Options()

    # 反检测参数
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    # 注意: 不要开启 headless
    # 不要加 --headless / --headless=new
    # CDE 的 WAF 会检测这些特征并返回空页面

    if chromedriver_path:
        service = Service(executable_path=chromedriver_path)
    else:
        service = Service()

    return webdriver.Chrome(service=service, options=options)


# ── 数据提取 ──────────────────────────────────────────────────

def extract_table_data(driver: webdriver.Chrome) -> list[dict]:
    """从 tbody#acceptVarietyInfoTbody 的 <td> 元素逐格提取数据。"""
    return driver.execute_script("""
        const tbody = document.getElementById('acceptVarietyInfoTbody');
        if (!tbody) return [];

        const columns = ['序号', '受理号', '药品名称', '药品类型',
                         '申请类型', '注册分类', '企业名称', '承办日期'];
        const result = [];

        for (const tr of tbody.querySelectorAll('tr')) {
            const tds = tr.querySelectorAll('td');
            if (tds.length === 0) continue;

            const row = {};
            tds.forEach((td, index) => {
                if (index < columns.length) {
                    row[columns[index]] = td.textContent.trim();
                }
            });
            result.push(row);
        }
        return result;
    """)


# ── Excel 导出 ──────────────────────────────────────────────────

def export_to_excel(records: list[dict], output_path: Path) -> None:
    """将记录导出为格式清晰的 Excel 表格。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "受理品种目录"

    # 样式定义
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写数据行
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 自动调整列宽
    for col_idx, col_name in enumerate(COLUMNS, 1):
        # 计算合适的列宽：取表头和数据中最大宽度
        max_width = len(col_name) * 2  # 中文字符占用约2个英文字符宽度
        for row_idx in range(2, len(records) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value or ""
            cell_width = len(str(cell_value)) * 1.3
            max_width = max(max_width, cell_width)
        # 限制最大宽度
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_width + 2, 60)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 设置行高
    ws.row_dimensions[1].height = 28

    wb.save(output_path)
    print(f"Excel 表格已保存 → {output_path}")


# ── 主流程 ────────────────────────────────────────────────────

def main() -> None:
    chromedriver_path = find_chromedriver()

    if chromedriver_path:
        print(f"ChromeDriver: {chromedriver_path}")
    else:
        print("错误: 找不到 chromedriver")
        print("")
        print("解决方法:")
        print("  1. brew install chromedriver  (Homebrew)")
        print("  2. 或手动下载: https://googlechromelabs.github.io/chrome-for-testing/")
        print("  3. 或指定环境变量: CHROMEDRIVER_PATH=/path/to/chromedriver python3 scrape_cde.py")
        sys.exit(1)

    print("正在启动浏览器（非 headless 模式）...")
    driver = create_driver(chromedriver_path)

    try:
        # ── 步骤 1: 打开页面 ──
        print(f"正在打开: {URL}")
        driver.get(URL)

        # ── 步骤 2: 等待 WAF 验证 ──
        # CDE 页面加载后，WAF 会插入 JS 挑战（加密计算）
        # 浏览器执行计算后返回真实页面。此过程不可见，需耐心等待。
        print("等待 WAF JavaScript 挑战通过...")
        start_time = time.time()

        # 轮询检测：等待 tbody 出现且包含数据行
        while time.time() - start_time < WAF_WAIT_SECONDS:
            try:
                tbody = driver.find_element(By.ID, "acceptVarietyInfoTbody")
                rows = tbody.find_elements(By.TAG_NAME, "tr")
                if rows:
                    elapsed = time.time() - start_time
                    print(f"  页面就绪（用时 {elapsed:.0f}s, 找到 {len(rows)} 行）")
                    break
            except Exception:
                pass
            time.sleep(1)

        # ── 步骤 3: 等待 LayUI 模板渲染完成 ──
        print("等待 LayUI 数据渲染...")
        try:
            WebDriverWait(driver, DATA_WAIT_SECONDS).until(
                lambda d: d.execute_script(
                    "const tb = document.getElementById('acceptVarietyInfoTbody');"
                    "const trs = tb ? tb.querySelectorAll('tr') : [];"
                    "return trs.length > 0 && trs[0].querySelectorAll('td').length >= 5;"
                )
            )
        except Exception:
            print("  警告: 数据渲染超时，尝试继续...")

        # ── 步骤 4: 提取数据 ──
        print("提取表格数据...")
        records = extract_table_data(driver)

        if not records:
            print("错误: 未提取到任何数据。可能原因:")
            print("  1. WAF 验证未通过——请观察浏览器窗口中是否显示了正常页面")
            print("  2. 网络延迟——适当增大 WAF_WAIT_SECONDS")
            print("  3. 页面结构变动——检查 tbody#acceptVarietyInfoTbody 是否存在")
            driver.save_screenshot(str(SCRIPT_DIR / "cde_debug.png"))
            print("  已保存调试截图到 cde_debug.png")
            sys.exit(1)

        # ── 步骤 5: 保存 ──
        with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)

        print(f"成功提取 {len(records)} 条记录 → {OUTPUT_PATH}")

        # Excel 导出
        export_to_excel(records, OUTPUT_XLSX_PATH)

        # ── 步骤 6: 打印结果 ──
        widths = [4, 16, 30, 16, 12, 8, 40, 12]
        sep = "=" * (sum(w + 1 for w in widths))
        print(f"\n{sep}")
        for h, w in zip(COLUMNS, widths):
            print(f"{h:<{w}}", end=" ")
        print(f"\n{sep}")
        for r in records:
            for h, w in zip(COLUMNS, widths):
                print(f"{r.get(h, ''):<{w}}", end=" ")
            print()
        print(sep)

    finally:
        driver.quit()
        print("\n浏览器已关闭。")


if __name__ == "__main__":
    main()
